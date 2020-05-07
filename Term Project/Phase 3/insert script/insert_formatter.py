import datetime
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from unicodedata import normalize

# get excel file name
def getFileName():
    fileName = input("what is the name of the .xlsx file?\n")
    return fileName

# get active worksheet
def getWS(fileName):
    source = (fileName + ".xlsx")
    wb = load_workbook(source)
    sheet = wb['Sheet1']
    return sheet

# get individual cell values in worksheet
def getContents(ws,colRange,rowRange):
    values = []
    #iterate thru sheet using col range
    for row in ws.iter_rows(min_row=1, max_col=colRange, max_row=rowRange, values_only=True):
        for cell in row:
            values.append(cell)
    return values


# get range of attributes and rows
def getRange():
    colRange = input("how many attributes?\n")
    rowRange = input("How many rows?\n")

    return int(colRange), int(rowRange)

def parseContents(data, colRange, rowRange):
    # parse data and separate new entry by every "colRange" attributes
    parsedData = []
    for value in data:
        # if value is not INT
        if not isinstance(value, int):
            # if value is VARCHAR
            if isinstance(value, str):
                new_str = normalize('NFKD', value)
                parsedData.append(new_str)

            # if table has DATE
            if isinstance(value, datetime.datetime):
                parsedData.append(value.strftime('%Y-%m-%d'))
        
            # if table has TIME           
            if isinstance(value, datetime.time):
                parsedData.append(value.strftime('%H:%M:%S'))
            
        else:
            parsedData.append(value)

    # make a new list of tuple sets with correct INSERT format
    compositeList = [tuple(parsedData[x:x+colRange]) for x in range(0, len(parsedData),colRange)]

    return compositeList

def displayOutput(compositeList, tableName):
    print("INSERT INTO " + tableName, compositeList[0])
    print("VALUES ")    
    for entry in range(1, len(compositeList)):
        if (entry == len(compositeList) - 1):
            print(compositeList[entry])
            print(";")
        else:
            print(str(compositeList[entry]) + ",")
    
    
def main():
    repeat = True
    while(repeat):
        path = os.path.dirname(os.path.realpath(__file__))  # Get the current working directory (cwd)
        os.chdir(path)
        cwd = os.getcwd()
        print(cwd)

        fileName = getFileName()
        ws = getWS(fileName)
        colRange,rowRange = getRange()
        data = getContents(ws, colRange, rowRange)
        compositeList = parseContents(data, colRange, rowRange)
        displayOutput(compositeList, fileName)
        user_input = input("format another table? y/n\n")
        if user_input == 'y':
            repeat = True
        else:
            repeat = False


if __name__ == "__main__":
    main()